from openpyxl.reader.excel import load_workbook

def readXlsxFile(path):
    #打开文件
    file = load_workbook(filename=path)
    #所有表格的名称
    #print(file.get_sheet_names())
    sheets = file.get_sheet_names()

    #拿出一个表格
    sheet = file.get_sheet_by_name(sheets[0])
    #最大行数
    print(sheet.max_row)
    #最大列数
    print(sheet.max_column)
    #表名
    print(sheet.title)

    #读取一张表的数据
    for lineNum in range(1,sheet.max_row + 1):
        print(lineNum)
        lineList = []
        for columnNum in range(1,sheet.max_column +1):
            #拿数据
            value = sheet.cell(row = lineNum,column=columnNum).value
            #if value != None:
            lineList.append(value)
        print(lineList)

path = r"E:\CX.PY\自动化办公和鼠标键盘模拟\Excel办公自动化\0.xlsx"
readXlsxFile(path)


