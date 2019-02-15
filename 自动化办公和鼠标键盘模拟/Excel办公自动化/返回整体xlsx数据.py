from openpyxl.reader.excel import load_workbook

def readXlsxFile(path):
    dic = {}
    file = load_workbook(filename=path)
    sheets = file.get_sheet_names()

    for sheetName in sheets:
        sheet = file.get_sheet_by_name(sheetName)
        #一张表所有数据
        sheetInfo = []
        for lineNum in range(1, sheet.max_row + 1):
            lineList = []
            for columnNum in range(1, sheet.max_column + 1):
                value = sheet.cell(row=lineNum, column=columnNum).value
                lineList.append(value)
            sheetInfo.append(lineList)

            #将一张表的数据存到字典
        dic[sheetName] = sheetInfo
    return dic




path = r"E:\CX.PY\自动化办公和鼠标键盘模拟\Excel办公自动化\0.xlsx"
dic = readXlsxFile(path)
print(dic)
print(len(dic))


